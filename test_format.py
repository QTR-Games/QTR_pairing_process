#!/usr/bin/env python3
"""
Simple test script to validate Excel format parsing
"""

import openpyxl
import pytest
from tkinter import filedialog
import tkinter as tk


pytestmark = pytest.mark.skip(reason="Manual interactive script; opens file chooser")


def simple_format_test():
    """Test Excel format parsing without database dependency"""
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Excel File to Analyze Format",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )

    if not file_path:
        print("No file selected")
        return

    try:
        print(f"Loading file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        if sheet is None:
            print("Error: No active sheet found")
            return

        print(f"Active sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print()

        friendly_team = sheet["A1"].value
        print(f"Friendly Team (A1): '{friendly_team}'")
        print()

        print("Analyzing team blocks with new logic:")
        team_blocks = []

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f"A{row}"].value
            if cell_value and str(cell_value).strip():
                team_name = str(cell_value).strip()

                header_row = row + 1
                if header_row <= sheet.max_row:
                    opponent_players = []
                    for col in ["C", "D", "E", "F", "G"]:
                        player_cell = sheet[f"{col}{header_row}"].value
                        if player_cell and str(player_cell).strip():
                            opponent_players.append(str(player_cell).strip())

                    if len(opponent_players) >= 3:
                        print(f"OK Found team block at row {row}: '{team_name}'")
                        print(f"  Opponent players: {opponent_players}")

                        ratings_data = []
                        data_row = row + 2
                        while data_row <= sheet.max_row:
                            friendly_player = sheet[f"B{data_row}"].value
                            if not friendly_player or not str(friendly_player).strip():
                                break

                            friendly_name = str(friendly_player).strip()

                            player_ratings = []
                            for col_idx, col in enumerate(["C", "D", "E", "F", "G"]):
                                if col_idx >= len(opponent_players):
                                    break
                                rating_cell = sheet[f"{col}{data_row}"].value
                                if rating_cell is not None:
                                    try:
                                        rating = int(rating_cell)
                                        player_ratings.append(rating)
                                    except Exception:
                                        player_ratings.append(3)
                                else:
                                    player_ratings.append(3)

                            ratings_data.append((friendly_name, player_ratings))
                            data_row += 1

                        print("  Friendly players and ratings:")
                        for fname, ratings in ratings_data:
                            print(f"    {fname}: {ratings}")

                        team_blocks.append((row, team_name))
                        print()

        print(f"Summary: Found {len(team_blocks)} valid team blocks")
        for row, name in team_blocks:
            print(f"  Row {row}: {name}")

        workbook.close()

    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    simple_format_test()
