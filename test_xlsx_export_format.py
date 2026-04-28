from types import SimpleNamespace

import openpyxl

from qtr_pairing_process.ui_manager_v2 import UiManager


class _Var:
    def __init__(self, value: str):
        self._value = value

    def get(self) -> str:
        return self._value


def test_export_xlsx_writes_import_ready_layout(monkeypatch, tmp_path):
    out_path = tmp_path / "import_ready.xlsx"

    info_calls = []
    error_calls = []

    monkeypatch.setattr(
        "qtr_pairing_process.ui_manager_v2.filedialog.asksaveasfilename",
        lambda **kwargs: str(out_path),
    )
    monkeypatch.setattr(
        "qtr_pairing_process.ui_manager_v2.messagebox.showinfo",
        lambda title, message: info_calls.append((title, message)),
    )
    monkeypatch.setattr(
        "qtr_pairing_process.ui_manager_v2.messagebox.showerror",
        lambda title, message: error_calls.append((title, message)),
    )

    rating_map = {
        ("F1", "O1"): 5,
        ("F1", "O2"): 4,
        ("F2", "O1"): 2,
    }

    dummy = SimpleNamespace(
        combobox_1=_Var("Friendly Team"),
        combobox_2=_Var("Opponent Team"),
        retrieve_team_data=lambda team_name: (
            team_name,
            ["F1", "F2", "F3", "F4", "F5"]
            if team_name == "Friendly Team"
            else ["O1", "O2", "O3", "O4", "O5"],
        ),
        db_manager=SimpleNamespace(query_team_id=lambda team_name: 1 if team_name == "Friendly Team" else 2),
        _fetch_matchup_ratings_by_name=lambda team_1_id, team_2_id, scenario_id=0: dict(rating_map),
        _operation_failed_error=lambda s: s,
        _operation_notice_info=lambda s: s,
        logger=SimpleNamespace(info=lambda *args, **kwargs: None, exception=lambda *args, **kwargs: None),
    )

    UiManager.export_xlsx(dummy)

    assert out_path.exists()
    assert not error_calls
    assert info_calls

    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook["Import Data"]

    assert sheet["A1"].value == "Friendly Team"
    assert sheet["A2"].value == "Opponent Team"

    assert [sheet.cell(row=2, column=col).value for col in range(3, 8)] == ["O1", "O2", "O3", "O4", "O5"]
    assert [sheet.cell(row=row, column=2).value for row in range(3, 8)] == ["F1", "F2", "F3", "F4", "F5"]

    assert sheet.cell(row=3, column=3).value == 5
    assert sheet.cell(row=3, column=4).value == 4
    assert sheet.cell(row=4, column=3).value == 2



def test_fetch_matchup_ratings_by_name_maps_rows():
    dummy = SimpleNamespace(
        db_manager=SimpleNamespace(
            query_sql=lambda sql: [
                ("F1", "O1", 5),
                ("F2", "O2", "3"),
                ("bad",),
            ]
        )
    )

    ratings = UiManager._fetch_matchup_ratings_by_name(dummy, 1, 2, scenario_id=0)

    assert ratings[("F1", "O1")] == 5
    assert ratings[("F2", "O2")] == 3
    assert len(ratings) == 2
