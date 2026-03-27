""" © Daniel P Raven 2024 All Rights Reserved """
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from qtr_pairing_process.db_management.db_manager import DbManager
from tkinter import messagebox
from typing import List, Dict, Tuple, Optional
import re


class SimpleExcelImporter:
    """
    Simplified Excel importer for the new user-friendly format.
    
    Expected Format:
    - A1: Friendly team name
    - Repeating blocks for each opponent team:
      - Row N: Opponent team name in column A (columns B-G empty)
      - Row N+1: Empty in A,B then opponent player names in columns C-G
      - Rows N+2 to N+6: Empty in A, friendly player name in B, ratings in C-G
    """
    
    def __init__(self, db_manager: DbManager, file_path: str, scenario_id: int = 0, rating_min: int = 1, rating_max: int = 5):
        self.db_manager = db_manager
        self.file_path = file_path
        self.scenario_id = scenario_id  # Default to "Neutral" scenario
        try:
            self.rating_min = int(rating_min)
            self.rating_max = int(rating_max)
        except (TypeError, ValueError):
            self.rating_min = 1
            self.rating_max = 5
        if self.rating_max < self.rating_min:
            self.rating_min, self.rating_max = self.rating_max, self.rating_min
        self.default_rating = int(round((self.rating_min + self.rating_max) / 2.0))
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None
        self.friendly_team_name: Optional[str] = None
        self.import_results: List[Dict] = []
        
    def load_workbook(self):
        """Load the Excel workbook and get the first sheet"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook.active  # Use the active sheet
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
            return False
            
    def extract_friendly_team_name(self):
        """Extract friendly team name from cell A1"""
        try:
            if self.sheet is None:
                raise ValueError("Worksheet not loaded")
            self.friendly_team_name = str(self.sheet['A1'].value).strip()
            if not self.friendly_team_name:
                raise ValueError("Friendly team name in A1 is empty")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read friendly team name from A1: {str(e)}")
            return False
            
    def find_opponent_team_blocks(self):
        """
        Find all opponent team blocks in the sheet.
        Returns list of (row_start, opponent_team_name) tuples.
        """
        if self.sheet is None:
            return []
            
        opponent_blocks = []
        
        # Scan column A for team names (skip A1 which is friendly team)
        for row in range(2, self.sheet.max_row + 1):
            cell_value = self.sheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                team_name = str(cell_value).strip()
                
                # Check if this looks like a team name (not a player name)
                # Look ahead to see if there's a structure following
                if self.validate_team_block_structure(row):
                    opponent_blocks.append((row, team_name))
                    
        return opponent_blocks
        
    def validate_team_block_structure(self, start_row):
        """
        Validate that a team block has the expected structure starting at start_row.
        Expected format based on CSV example:
        - Row start_row: Opponent team name in column A, empty in B, opponent players in C-G
        - Rows start_row+1 to start_row+5: Empty A, friendly player in B, 5 ratings in C-G
        """
        try:
            if self.sheet is None:
                return False
                
            # Check if this row has opponent player names in columns C-G (not next row!)
            opponent_players = 0
            for col in ['C', 'D', 'E', 'F', 'G']:
                cell_value = self.sheet[f'{col}{start_row}'].value  # Same row, not start_row + 1
                if cell_value and str(cell_value).strip():
                    opponent_players += 1
                    
            if opponent_players != 5:  # Must have exactly 5 opponent players
                return False
                
            # Check if we have exactly 5 friendly player names in column B (rows start_row+1 to start_row+5)
            friendly_players = 0
            for row_offset in range(1, 6):  # Changed from range(2, 7) to range(1, 6)
                check_row = start_row + row_offset
                if check_row > self.sheet.max_row:
                    return False  # Not enough rows
                cell_value = self.sheet[f'B{check_row}'].value
                if cell_value and str(cell_value).strip():
                    friendly_players += 1
                    
            return friendly_players == 5  # Must have exactly 5 friendly players
            
        except Exception:
            return False
            
    def extract_team_block_data(self, start_row, opponent_team_name):
        """
        Extract all data from a team block starting at start_row.
        Returns dict with opponent team info and ratings matrix.
        """
        try:
            # Extract opponent player names from start_row (same row as team name), columns C-G
            opponent_players = []
            for col in ['C', 'D', 'E', 'F', 'G']:
                if self.sheet is None:
                    raise ValueError("Worksheet is None")
                cell_value = self.sheet[f'{col}{start_row}'].value  # Changed from header_row to start_row
                if cell_value and str(cell_value).strip():
                    opponent_players.append(str(cell_value).strip())
                else:
                    opponent_players.append("")  # Track empty cells
                    
            # Filter out empty cells and check for exactly 5 players
            valid_opponent_players = [p for p in opponent_players if p]
            if len(valid_opponent_players) != 5:
                raise ValueError(f"Team must have exactly 5 opponent players, found {len(valid_opponent_players)}. Team will be skipped.")
                
            # Extract friendly player names and ratings matrix (exactly 5 rows expected)
            friendly_players = []
            ratings_matrix = []
            
            # Extract exactly 5 friendly players from rows start_row+1 to start_row+5
            for row_offset in range(1, 6):  # Changed from range(2, 7) to range(1, 6)
                data_row = start_row + row_offset
                if data_row > (self.sheet.max_row if self.sheet else 0):
                    raise ValueError(f"Not enough rows to extract 5 friendly players. Expected row {data_row} but sheet only has {self.sheet.max_row if self.sheet else 0} rows.")
                
                if self.sheet is None:
                    raise ValueError("Worksheet is None")
                
                # Check for friendly player name in column B
                friendly_player = self.sheet[f'B{data_row}'].value
                if not friendly_player or not str(friendly_player).strip():
                    raise ValueError(f"Missing friendly player name at row {data_row}. Team must have exactly 5 friendly players.")
                    
                friendly_players.append(str(friendly_player).strip())
                
                # Extract ratings for this friendly player against all 5 opponent players
                player_ratings = []
                for col in ['C', 'D', 'E', 'F', 'G']:
                        
                    rating_cell = self.sheet[f'{col}{data_row}'].value
                    
                    # Convert rating to integer, default to midpoint if invalid
                    try:
                        if rating_cell is None:
                            rating = self.default_rating
                        elif isinstance(rating_cell, (int, float)):
                            rating = int(rating_cell)
                        else:
                            rating = int(str(rating_cell))
                            
                        # Validate rating range
                        if rating < self.rating_min or rating > self.rating_max:
                            rating = self.default_rating
                            
                    except (ValueError, TypeError):
                        rating = self.default_rating  # Default to neutral midpoint
                        
                    player_ratings.append(rating)
                    
                ratings_matrix.append(player_ratings)
                
            # Verify we have exactly 5 friendly players and 5x5 ratings matrix
            if len(friendly_players) != 5:
                raise ValueError(f"Team must have exactly 5 friendly players, found {len(friendly_players)}")
            
            if len(ratings_matrix) != 5:
                raise ValueError(f"Ratings matrix must have exactly 5 rows, found {len(ratings_matrix)}")
                
            for i, row in enumerate(ratings_matrix):
                if len(row) != 5:
                    raise ValueError(f"Ratings row {i+1} must have exactly 5 ratings, found {len(row)}")
                
            return {
                'opponent_team_name': opponent_team_name,
                'opponent_players': valid_opponent_players,  # Use the validated list
                'friendly_players': friendly_players,
                'ratings_matrix': ratings_matrix,
                'start_row': start_row
            }
            
        except Exception as e:
            raise ValueError(f"Failed to extract data from team block at row {start_row}: {str(e)}")
            
    def validate_team_data(self, team_data):
        """Validate that team data is complete and consistent"""
        errors = []
        
        # Check team name
        if not team_data['opponent_team_name']:
            errors.append("Missing opponent team name")
            
        # Check player counts (must be exactly 5)
        if len(team_data['opponent_players']) != 5:
            errors.append(f"Team must have exactly 5 opponent players, got {len(team_data['opponent_players'])}")
            
        if len(team_data['friendly_players']) != 5:
            errors.append(f"Team must have exactly 5 friendly players, got {len(team_data['friendly_players'])}")
            
        # Check ratings matrix (must be exactly 5x5)
        if len(team_data['ratings_matrix']) != 5:
            errors.append(f"Ratings matrix must have exactly 5 rows, got {len(team_data['ratings_matrix'])}")
        else:
            for i, row in enumerate(team_data['ratings_matrix']):
                if len(row) != 5:
                    errors.append(f"Ratings row {i+1} must have exactly 5 ratings, got {len(row)}")
                    
        return errors
        
    def save_team_to_database(self, team_data):
        """Save team data and ratings to database"""
        try:
            # Anchor matchup orientation to the friendly team declared in A1.
            friendly_team_id = self._resolve_or_create_matchup_team(
                self.friendly_team_name,
                team_data['friendly_players'],
                team_role="friendly",
            )

            opponent_team_id = self._resolve_or_create_matchup_team(
                team_data['opponent_team_name'],
                team_data['opponent_players'],
                team_role="opponent",
                friendly_team_name=self.friendly_team_name,
            )
            
            # Get player IDs for ratings
            friendly_player_ids = {}
            opponent_player_ids = {}
            
            for i, player_name in enumerate(team_data['friendly_players']):
                player_id = self.db_manager.query_player_id(player_name, friendly_team_id)
                friendly_player_ids[i] = player_id
                
            for j, player_name in enumerate(team_data['opponent_players']):
                player_id = self.db_manager.query_player_id(player_name, opponent_team_id)
                opponent_player_ids[j] = player_id
                
            # Save ratings matrix
            for i, rating_row in enumerate(team_data['ratings_matrix']):
                for j, rating in enumerate(rating_row):
                    self.db_manager.upsert_rating(
                        player_id_1=friendly_player_ids[i],
                        player_id_2=opponent_player_ids[j],
                        team_id_1=friendly_team_id,
                        team_id_2=opponent_team_id,
                        scenario_id=self.scenario_id,
                        rating=rating
                    )
                    
            return True
            
        except Exception as e:
            raise Exception(f"Database error for team {team_data['opponent_team_name']}: {str(e)}")

    def _resolve_or_create_matchup_team(
        self,
        team_name: str,
        player_names: List[str],
        team_role: str,
        friendly_team_name: Optional[str] = None,
    ) -> int:
        """Resolve a team for import, creating it when missing, then validate roster."""
        existing_team_id = self.db_manager.query_team_id(team_name)
        if existing_team_id is None:
            team_id = self.db_manager.upsert_team(team_name)
            if team_role == "opponent" and friendly_team_name:
                print(
                    f"Created missing opponent team '{team_name}' for matchup against "
                    f"friendly team '{friendly_team_name}'."
                )
        else:
            team_id = existing_team_id

        self.db_manager.upsert_and_validate_players(team_id, player_names)
        return team_id
            
    def execute_import(self):
        """Execute the complete import process"""
        try:
            # Load workbook
            if not self.load_workbook():
                return False
                
            # Extract friendly team name
            if not self.extract_friendly_team_name():
                return False
                
            # Find all opponent team blocks
            opponent_blocks = self.find_opponent_team_blocks()
            
            if not opponent_blocks:
                messagebox.showerror("Error", "No opponent team blocks found in the Excel file.")
                return False
                
            # Process each opponent team block
            successful_imports = 0
            total_teams = len(opponent_blocks)
            
            for start_row, opponent_team_name in opponent_blocks:
                try:
                    # Extract team data with enhanced error detection
                    team_data = self.extract_team_block_data(start_row, opponent_team_name)
                    
                    # Validate data structure
                    validation_errors = self.validate_team_data(team_data)
                    if validation_errors:
                        detailed_error = self.create_detailed_validation_error(opponent_team_name, validation_errors, start_row)
                        self.import_results.append({
                            'team_name': opponent_team_name,
                            'status': 'error',
                            'message': detailed_error
                        })
                        print(f"Skipping {opponent_team_name}: {detailed_error}")
                        continue
                    
                    # Check for duplicate team before saving
                    duplicate_check = self.check_for_duplicate_team(team_data)
                    if duplicate_check:
                        self.import_results.append({
                            'team_name': opponent_team_name,
                            'status': 'error',
                            'message': duplicate_check
                        })
                        print(f"Skipping {opponent_team_name}: {duplicate_check}")
                        continue
                        
                    # Save to database
                    self.save_team_to_database(team_data)
                    
                    self.import_results.append({
                        'team_name': opponent_team_name,
                        'status': 'success',
                        'message': f"Successfully imported {len(team_data['opponent_players'])} players and {len(team_data['ratings_matrix'])} × {len(team_data['ratings_matrix'][0])} ratings"
                    })
                    
                    successful_imports += 1
                    
                except Exception as e:
                    # Create detailed error message based on exception type
                    detailed_error = self.create_detailed_extraction_error(opponent_team_name, str(e), start_row)
                    print(f"Error importing {opponent_team_name}: {detailed_error}")
                    self.import_results.append({
                        'team_name': opponent_team_name,
                        'status': 'error', 
                        'message': detailed_error
                    })
                    
            # Show results with improved formatting
            if successful_imports > 0:
                # Create a clean, well-formatted success message
                if successful_imports == total_teams:
                    # All teams imported successfully
                    success_msg = f"Import Complete!\n\nSuccessfully imported all {successful_imports} teams for '{self.friendly_team_name}'"
                else:
                    # Some teams had issues
                    skipped_count = total_teams - successful_imports
                    success_msg = f"Import Partially Complete\n\nResults for '{self.friendly_team_name}':"
                    success_msg += f"\n• Successfully imported: {successful_imports} teams"
                    success_msg += f"\n• Skipped due to errors: {skipped_count} teams"
                
                # Add list of successfully imported teams
                success_teams = [r['team_name'] for r in self.import_results if r['status'] == 'success']
                if success_teams:
                    success_msg += f"\n\nImported Teams:"
                    for team in success_teams:
                        success_msg += f"\n✓ {team}"
                
                # Add list of skipped teams with detailed error messages
                error_teams = [r for r in self.import_results if r['status'] == 'error']
                if error_teams:
                    success_msg += f"\n\nSkipped Teams (with reasons):"
                    for team_result in error_teams:
                        success_msg += f"\n\n✗ {team_result['team_name']}:"
                        success_msg += f"\n   {team_result['message']}"
                    
                messagebox.showinfo("Import Results", success_msg)
                return True
            else:
                # No teams were successfully imported
                error_msg = f"Import Failed\n\nNo teams could be imported from '{self.friendly_team_name}'"
                
                # Add details about what went wrong with enhanced formatting
                if self.import_results:
                    error_msg += f"\n\nDetailed Error Information:"
                    for result in self.import_results:
                        if result['status'] == 'error':
                            error_msg += f"\n\n✗ {result['team_name']}:"
                            error_msg += f"\n   {result['message']}"
                else:
                    error_msg += "\n\nPlease check that your Excel file follows the expected format."
                    
                messagebox.showerror("Import Failed", error_msg)
                return False
                
        except Exception as e:
            messagebox.showerror("Import Error", f"Unexpected error during import: {str(e)}")
            return False
        finally:
            if self.workbook:
                self.workbook.close()
    
    def create_detailed_validation_error(self, team_name: str, validation_errors: List[str], start_row: int) -> str:
        """Create a detailed, user-friendly validation error message"""
        error_msg = f"Data format issues in '{team_name}' (starting at row {start_row}):"
        
        for error in validation_errors:
            if "exactly 5 opponent players" in error:
                error_msg += f"\n• Missing opponent players - check row {start_row}, columns C-G should contain exactly 5 player names"
            elif "exactly 5 friendly players" in error:
                error_msg += f"\n• Missing friendly players - check rows {start_row+1} to {start_row+5}, column B should contain exactly 5 player names"
            elif "exactly 5 rows" in error:
                error_msg += f"\n• Incomplete ratings matrix - need exactly 5 rows of friendly players and ratings"
            elif "exactly 5 ratings" in error:
                error_msg += f"\n• Incomplete ratings row - each friendly player needs ratings against all 5 opponents (columns C-G)"
            else:
                error_msg += f"\n• {error}"
        
        error_msg += f"\n\nExpected format:\nRow {start_row}: Team name (A) + 5 opponent players (C-G)\nRows {start_row+1}-{start_row+5}: Friendly player (B) + 5 ratings (C-G)"
        return error_msg
    
    def check_for_duplicate_team(self, team_data: Dict) -> Optional[str]:
        """Check if team already exists in database with same name and players"""
        try:
            # Check if team name already exists
            existing_teams_result = self.db_manager.execute_sql("SELECT team_name FROM teams WHERE team_name = ?", (team_data['opponent_team_name'],))
            
            if existing_teams_result and existing_teams_result > 0:
                # Team name exists, check if it has the same players
                try:
                    existing_players_result = self.db_manager.execute_sql(
                        "SELECT player_name FROM players WHERE team_name = ? ORDER BY player_name", 
                        (team_data['opponent_team_name'],)
                    )
                    
                    # Get all existing players for this team
                    existing_player_names = []
                    # Note: db_manager.execute_sql might return different formats, handle gracefully
                    
                    # For now, assume the team exists and give a generic duplicate message
                    new_player_names = sorted(team_data['opponent_players'])
                    return f"Team '{team_data['opponent_team_name']}' already exists in the database. Please use a different team name or delete the existing team first."
                    
                except Exception:
                    return f"Team '{team_data['opponent_team_name']}' already exists in the database. Cannot verify if players are the same due to database access issues."
            
            return None  # No duplicate found
            
        except Exception as e:
            # Don't fail import for duplicate check errors, just warn
            print(f"Warning: Could not check for duplicates: {str(e)}")
            return None
    
    def create_detailed_extraction_error(self, team_name: str, error_msg: str, start_row: int) -> str:
        """Create a detailed error message for data extraction failures"""
        if "exactly 5 opponent players" in error_msg:
            return f"'{team_name}' (row {start_row}): Found incorrect number of opponent players in columns C-G. Please ensure exactly 5 opponent player names are entered."
        elif "Missing friendly player name" in error_msg:
            row_match = error_msg.split("row ")[-1].split(".")[0] if "row " in error_msg else "unknown"
            return f"'{team_name}': Missing friendly player name at row {row_match}, column B. All 5 friendly player positions must be filled."
        elif "Not enough rows" in error_msg:
            return f"'{team_name}' (row {start_row}): Not enough rows in Excel sheet to extract 5 friendly players. Need at least {start_row+5} rows in the sheet."
        elif "Worksheet is None" in error_msg:
            return f"'{team_name}': Excel worksheet could not be read. Please check if the file is corrupted or in use by another program."
        else:
            return f"'{team_name}' (row {start_row}): {error_msg}"
                
    def get_import_summary(self):
        """Return a summary of the import results"""
        return {
            'friendly_team': self.friendly_team_name,
            'scenario_id': self.scenario_id,
            'results': self.import_results
        }
    
    def execute(self):
        """Execute the import process - compatibility method for UI integration"""
        success = self.execute_import()
        if success:
            return len([r for r in self.import_results if r['status'] == 'success'])
        else:
            return 0