"""Simple XLSX exporter compatible with SimpleExcelImporter format."""

from __future__ import annotations

from typing import List

import openpyxl


class SimpleExcelExporter:
    """Export one friendly-vs-opponent block in the active simple XLSX template.

    Template emitted:
    - A1: friendly team name
    - Row 2: A=opponent team name, C:G=opponent players
    - Rows 3-7: B=friendly player name, C:G=ratings
    """

    def __init__(
        self,
        file_path: str,
        friendly_team_name: str,
        opponent_team_name: str,
        friendly_players: List[str],
        opponent_players: List[str],
        ratings_matrix: List[List[int]],
    ) -> None:
        self.file_path = file_path
        self.friendly_team_name = (friendly_team_name or "").strip()
        self.opponent_team_name = (opponent_team_name or "").strip()
        self.friendly_players = [str(name).strip() for name in friendly_players]
        self.opponent_players = [str(name).strip() for name in opponent_players]
        self.ratings_matrix = ratings_matrix

    def _validate(self) -> None:
        if not self.file_path:
            raise ValueError("File path is required")
        if not self.friendly_team_name:
            raise ValueError("Friendly team name is required")
        if not self.opponent_team_name:
            raise ValueError("Opponent team name is required")
        if len(self.friendly_players) != 5:
            raise ValueError(f"Friendly team must have exactly 5 players, got {len(self.friendly_players)}")
        if len(self.opponent_players) != 5:
            raise ValueError(f"Opponent team must have exactly 5 players, got {len(self.opponent_players)}")
        if len(self.ratings_matrix) != 5:
            raise ValueError(f"Ratings matrix must have exactly 5 rows, got {len(self.ratings_matrix)}")

        for idx, row in enumerate(self.ratings_matrix):
            if len(row) != 5:
                raise ValueError(f"Ratings row {idx + 1} must have exactly 5 values, got {len(row)}")

    def execute(self) -> str:
        self._validate()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Data"

        # A1: friendly team name
        ws["A1"] = self.friendly_team_name

        # Row 2: opponent team block header
        ws["A2"] = self.opponent_team_name
        for col_idx, opponent_name in enumerate(self.opponent_players, start=3):
            ws.cell(row=2, column=col_idx, value=opponent_name)

        # Rows 3..7: friendly players + 5x5 ratings
        for row_offset, friendly_name in enumerate(self.friendly_players, start=1):
            excel_row = 2 + row_offset
            ws.cell(row=excel_row, column=2, value=friendly_name)
            ratings_row = self.ratings_matrix[row_offset - 1]
            for col_offset, rating in enumerate(ratings_row, start=1):
                ws.cell(row=excel_row, column=2 + col_offset, value=int(rating))

        wb.save(self.file_path)
        wb.close()
        return self.file_path
