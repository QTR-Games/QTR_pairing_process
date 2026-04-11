#!/usr/bin/env python3
"""
Test the 5-player enforcement in SimpleExcelImporter
"""

import openpyxl
import pytest
from tkinter import filedialog
import tkinter as tk


pytestmark = pytest.mark.skip(reason="Manual interactive script; opens file chooser")


def test_five_player_validation():
    """Test that the importer correctly validates 5-player teams"""
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Excel File to Test 5-Player Validation",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )

    if not file_path:
        print("No file selected")
        return

    try:
        print(f"Testing 5-player validation on: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        if sheet is None:
            print("Error: No active sheet found")
            return

        friendly_team = sheet["A1"].value
        print(f"Friendly Team: '{friendly_team}'")
        print()

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f"A{row}"].value
            if cell_value and str(cell_value).strip():
                team_name = str(cell_value).strip()
                print(f"Testing team block at row {row}: '{team_name}'")

                header_row = row + 1
                if header_row <= sheet.max_row:
                    opponent_players = []
                    for col in ["C", "D", "E", "F", "G"]:
                        player_cell = sheet[f"{col}{header_row}"].value
                        if player_cell and str(player_cell).strip():
                            opponent_players.append(str(player_cell).strip())

                    print(f"  Opponent players found: {len(opponent_players)}")
                    if len(opponent_players) == 5:
                        print(f"  OK Opponent players: {opponent_players}")
                    else:
                        print(
                            f"  ERROR Expected exactly 5 opponent players, found {len(opponent_players)}: {opponent_players}"
                        )
                        print("  -> This team would be SKIPPED during import")
                        continue

                    friendly_players = []
                    for row_offset in range(2, 7):
                        check_row = row + row_offset
                        if check_row <= sheet.max_row:
                            friendly_player = sheet[f"B{check_row}"].value
                            if friendly_player and str(friendly_player).strip():
                                friendly_players.append(str(friendly_player).strip())

                    print(f"  Friendly players found: {len(friendly_players)}")
                    if len(friendly_players) == 5:
                        print(f"  OK Friendly players: {friendly_players}")
                    else:
                        print(
                            f"  ERROR Expected exactly 5 friendly players, found {len(friendly_players)}: {friendly_players}"
                        )
                        print("  -> This team would be SKIPPED during import")
                        continue

                    ratings_valid = True
                    for row_offset in range(2, 7):
                        check_row = row + row_offset
                        if check_row <= sheet.max_row:
                            row_ratings = []
                            for col in ["C", "D", "E", "F", "G"]:
                                rating_cell = sheet[f"{col}{check_row}"].value
                                if rating_cell is not None:
                                    try:
                                        rating = int(rating_cell)
                                        if 1 <= rating <= 5:
                                            row_ratings.append(rating)
                                        else:
                                            row_ratings.append(3)
                                    except Exception:
                                        row_ratings.append(3)
                                else:
                                    row_ratings.append(3)

                            if len(row_ratings) != 5:
                                ratings_valid = False
                                break

                    if ratings_valid:
                        print("  OK Ratings matrix is valid (5x5)")
                        print("  -> This team would be SUCCESSFULLY IMPORTED")
                    else:
                        print("  ERROR Ratings matrix is invalid")
                        print("  -> This team would be SKIPPED during import")

                print()

        workbook.close()

    except Exception as e:
        print(f"Error testing validation: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    test_five_player_validation()
