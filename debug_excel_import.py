#!/usr/bin/env python3
"""
Debug script to examine Excel file structure and troubleshoot import issues
"""

import openpyxl
from tkinter import filedialog
import tkinter as tk

def debug_excel_file():
    """Debug an Excel file to understand its structure"""
    # Create a simple root window for the file dialog
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Ask user to select Excel file
    file_path = filedialog.askopenfilename(
        title="Select Excel File to Debug",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    if not file_path:
        print("No file selected")
        return
        
    try:
        # Load workbook
        print(f"Loading file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        if sheet is None:
            print("Error: No active sheet found")
            return
            
        print(f"Active sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print(f"Max column: {sheet.max_column}")
        print()
        
        # Show cell A1 (friendly team)
        a1_value = sheet['A1'].value
        print(f"A1 (Friendly Team): '{a1_value}'")
        print()
        
        # Scan first 20 rows to understand structure
        print("Structure analysis (first 20 rows):")
        print("Row | A        | B        | C        | D        | E        | F        | G")
        print("----|----------|----------|----------|----------|----------|----------|----------")
        
        for row in range(1, min(21, sheet.max_row + 1)):
            row_data = []
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                cell_value = sheet[f'{col}{row}'].value
                if cell_value is None:
                    display_value = ""
                else:
                    display_value = str(cell_value)[:8]  # Truncate for display
                row_data.append(f"{display_value:8}")
            
            print(f"{row:3} | " + " | ".join(row_data))
        
        print()
        
        # Look for potential team blocks
        print("Potential team blocks found:")
        team_blocks = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                team_name = str(cell_value).strip()
                print(f"Row {row}: '{team_name}'")
                
                # Check if next row has player names in C-G
                player_row = row + 1
                players_found = []
                for col in ['C', 'D', 'E', 'F', 'G']:
                    player_cell = sheet[f'{col}{player_row}'].value
                    if player_cell and str(player_cell).strip():
                        players_found.append(str(player_cell).strip())
                
                if players_found:
                    print(f"  -> Players in row {player_row}: {players_found}")
                    
                    # Check for friendly players and ratings
                    friendly_players = []
                    for row_offset in range(2, 7):
                        check_row = row + row_offset
                        if check_row > sheet.max_row:
                            break
                        friendly_player = sheet[f'B{check_row}'].value
                        if friendly_player and str(friendly_player).strip():
                            friendly_players.append(str(friendly_player).strip())
                    
                    if friendly_players:
                        print(f"  -> Friendly players: {friendly_players}")
                        team_blocks.append((row, team_name))
                    else:
                        print(f"  -> No friendly players found")
                else:
                    print(f"  -> No opponent players found in row {player_row}")
                print()
        
        print(f"Valid team blocks identified: {len(team_blocks)}")
        for row, name in team_blocks:
            print(f"  Row {row}: {name}")
            
        workbook.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_excel_file()