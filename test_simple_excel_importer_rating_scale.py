from pathlib import Path

import openpyxl

from qtr_pairing_process.excel_management.simple_excel_importer import SimpleExcelImporter


class _DummyDb:
    pass


def _build_import_ready_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Friendly"

    # Opponent block start at row 2
    ws["A2"] = "Opponent"
    ws["C2"] = "O1"
    ws["D2"] = "O2"
    ws["E2"] = "O3"
    ws["F2"] = "O4"
    ws["G2"] = "O5"

    friendly_players = ["F1", "F2", "F3", "F4", "F5"]
    ratings_rows = [
        [10, 9, 8, 7, 6],
        [9, 8, 7, 6, 5],
        [8, 7, 6, 5, 4],
        [7, 6, 5, 4, 3],
        [6, 5, 4, 3, 2],
    ]

    for row_offset, player_name in enumerate(friendly_players, start=1):
        row = 2 + row_offset
        ws.cell(row=row, column=2, value=player_name)
        for col_offset, rating in enumerate(ratings_rows[row_offset - 1], start=3):
            ws.cell(row=row, column=col_offset, value=rating)

    wb.save(path)
    wb.close()


def test_simple_excel_importer_preserves_1_to_10_scale(tmp_path):
    file_path = tmp_path / "import_ready_1_10.xlsx"
    _build_import_ready_workbook(file_path)

    importer = SimpleExcelImporter(
        db_manager=_DummyDb(),
        file_path=str(file_path),
        scenario_id=0,
        rating_min=1,
        rating_max=10,
    )

    assert importer.load_workbook() is True
    assert importer.extract_friendly_team_name() is True

    blocks = importer.find_opponent_team_blocks()
    assert len(blocks) == 1

    start_row, opponent_team_name = blocks[0]
    team_data = importer.extract_team_block_data(start_row, opponent_team_name)

    assert team_data["ratings_matrix"][0][0] == 10
    assert team_data["ratings_matrix"][0][1] == 9
    assert team_data["ratings_matrix"][4][4] == 2
