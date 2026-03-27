from pathlib import Path

import openpyxl
import pytest

from qtr_pairing_process.ui_manager_v2 import UiManager


COLUMNS = [
    "schema_version",
    "app_export_version",
    "source_db_fingerprint",
    "source_roster_hash",
    "source_team_id",
    "source_team_name",
    "source_player_id",
    "source_player_name",
    "opponent_team_id",
    "opponent_team_name",
    "opponent_player_id",
    "opponent_player_name",
    "scenario_id",
    "rating",
    "comment",
]


def _canonical_rows():
    base = {
        "schema_version": "player_ratings_export_v1",
        "app_export_version": "1.0",
        "source_db_fingerprint": "db-fp",
        "source_roster_hash": "roster-hash",
        "source_team_id": "1",
        "source_team_name": "Friendly",
        "source_player_id": "11",
        "source_player_name": "F1",
        "comment": "",
    }
    row1 = dict(base)
    row1.update(
        {
            "opponent_team_id": "2",
            "opponent_team_name": "Enemy",
            "opponent_player_id": "21",
            "opponent_player_name": "E1",
            "scenario_id": "0",
            "rating": "4",
            "comment": "note",
        }
    )
    row2 = dict(base)
    row2.update(
        {
            "opponent_team_id": "2",
            "opponent_team_name": "Enemy",
            "opponent_player_id": "22",
            "opponent_player_name": "E2",
            "scenario_id": "1",
            "rating": "3",
        }
    )
    return [row1, row2]


class _DbStub:
    def get_db_fingerprint(self):
        return "different-fp"

    def get_team_roster_hash(self, _team_id):
        return "different-roster"

    def query_team_id(self, team_name):
        return 2 if team_name == "Enemy" else None

    def query_player_id(self, player_name, team_id):
        return 20 if team_id == 2 and player_name == "E1" else None

    def query_sql_params(self, _sql, _params):
        return [(10,)]


class _ImportUiDummy:
    def __init__(self, db_manager=None):
        self.current_rating_system = "1-5"
        self.db_manager = db_manager

    def _individual_player_export_columns(self):
        return list(COLUMNS)

    def _schema_missing_columns_error(self, missing_columns):
        return UiManager._schema_missing_columns_error(self, missing_columns)

    def _schema_version_error(self, schema_version):
        return UiManager._schema_version_error(self, schema_version)

    def _identity_mismatch_error(self, details):
        return UiManager._identity_mismatch_error(self, details)

    def _identity_resolution_error(self, details):
        return UiManager._identity_resolution_error(self, details)

    def _lineage_fallback_warning_text(self):
        return UiManager._lineage_fallback_warning_text(self)

    def _partial_export_warning_text(self, actual_rows, expected_rows):
        return UiManager._partial_export_warning_text(self, actual_rows, expected_rows)

    def _format_individual_import_validation_errors(self, errors):
        return UiManager._format_individual_import_validation_errors(self, errors)

    def _load_individual_player_csv_rows(self, file_path):
        return UiManager._load_individual_player_csv_rows(self, file_path)

    def _load_individual_player_xlsx_rows(self, file_path):
        return UiManager._load_individual_player_xlsx_rows(self, file_path)

    def _load_individual_player_rows(self, file_path):
        return UiManager._load_individual_player_rows(self, file_path)


def _write_csv(file_path: Path, rows):
    import csv

    with open(file_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_xlsx(file_path: Path, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
    workbook.save(file_path)
    workbook.close()


def test_csv_and_xlsx_load_to_same_rows(tmp_path):
    rows = _canonical_rows()
    csv_path = tmp_path / "player_rows.csv"
    xlsx_path = tmp_path / "player_rows.xlsx"

    _write_csv(csv_path, rows)
    _write_xlsx(xlsx_path, rows)

    ui = _ImportUiDummy()

    csv_rows = UiManager._load_individual_player_rows(ui, str(csv_path))
    xlsx_rows = UiManager._load_individual_player_rows(ui, str(xlsx_path))

    assert csv_rows == xlsx_rows


def test_validate_collects_multiple_row_column_errors():
    first = _canonical_rows()[0]
    bad_rows = [dict(first), dict(first), dict(first)]

    bad_rows[0]["scenario_id"] = "not-an-int"
    bad_rows[1]["rating"] = "999"
    bad_rows[1]["opponent_player_name"] = "E1"
    bad_rows[2]["comment"] = "x" * 2001

    ui = _ImportUiDummy(db_manager=_DbStub())

    with pytest.raises(ValueError) as exc:
        UiManager._validate_individual_import_rows(ui, bad_rows, 1, "Friendly", 11, "F1")

    message = str(exc.value)
    assert "row 2, column 'scenario_id'" in message
    assert "row 3, column 'rating'" in message
    assert "row 4, column 'comment'" in message
